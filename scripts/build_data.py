#!/usr/bin/env python3
"""資源エネルギー庁「石油製品小売市況調査」時系列Excel(s5)をJSONに変換する。

入力: pl007/xlsx/{yymmdd}s5.xlsx（シート: ハイオク/レギュラー/軽油/灯油店頭/灯油配達）
出力:
  data/latest.json           最新週+前週の全国・都道府県別価格（4油種）
  data/history/national.json 全国平均の全履歴（4油種）
  data/history/{NN}.json     都道府県別の全履歴（NN=01..47）
  data/meta.json             生成日時・調査日・出典

灯油は円/18Lで公表されるため円/Lに正規化する。
"""
from __future__ import annotations

import datetime as dt
import json
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"

# シート名 -> (油種キー, 円/Lへの除数)
SHEETS = {
    "ハイオク": ("high_octane", 1),
    "レギュラー": ("regular", 1),
    "軽油": ("diesel", 1),
    "灯油店頭": ("kerosene", 18),
}

PREF_NAMES = [
    "北海道", "青森", "岩手", "宮城", "秋田", "山形", "福島",
    "茨城", "栃木", "群馬", "埼玉", "千葉", "東京", "神奈川",
    "新潟", "富山", "石川", "福井", "山梨", "長野", "岐阜",
    "静岡", "愛知", "三重", "滋賀", "京都", "大阪", "兵庫",
    "奈良", "和歌山", "鳥取", "島根", "岡山", "広島", "山口",
    "徳島", "香川", "愛媛", "高知", "福岡", "佐賀", "長崎",
    "熊本", "大分", "宮崎", "鹿児島", "沖縄",
]
PREF_CODE = {name: f"{i + 1:02d}" for i, name in enumerate(PREF_NAMES)}


def normalize_header(value: object) -> str:
    """ヘッダーセルの空白ゆれ（全角空白・改行）を除去して県名に寄せる。"""
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value))
    return re.sub(r"\s+", "", text)


def header_to_code(header: str) -> str | None:
    """ヘッダー文字列を県コード('01'-'47')か全国('00')に変換。局集計列はNone。"""
    if header == "全国":
        return "00"
    if header in ("北海道", "北海道局"):
        return "01"
    if header in ("沖縄", "沖縄局"):
        return "47"
    if header.endswith("局"):  # 東北局・関東局などの地域集計は除外
        return None
    return PREF_CODE.get(header)


def parse_sheet(ws) -> tuple[list[str | None], list[tuple[dt.date, list]]]:
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    codes = [header_to_code(normalize_header(h)) for h in header]
    series: list[tuple[dt.date, list]] = []
    for row in rows:
        date = row[1] if len(row) > 1 else None
        if isinstance(date, dt.datetime):
            series.append((date.date(), list(row)))
    return codes, series


def to_price(value: object, divisor: int) -> float | None:
    if not isinstance(value, (int, float)):
        return None
    return round(float(value) / divisor, 1)


def build(xlsx_path: Path) -> None:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)

    # fuel -> code -> [(date, price)]
    history: dict[str, dict[str, list[tuple[str, float]]]] = {}
    latest: dict[str, dict] = {}

    for sheet_name, (fuel, divisor) in SHEETS.items():
        if sheet_name not in wb.sheetnames:
            print(f"WARN: sheet {sheet_name} not found", file=sys.stderr)
            continue
        codes, series = parse_sheet(wb[sheet_name])
        fuel_hist: dict[str, list[tuple[str, float]]] = {}
        for date, row in series:
            iso = date.isoformat()
            for idx, code in enumerate(codes):
                if code is None or idx >= len(row):
                    continue
                price = to_price(row[idx], divisor)
                if price is None:
                    continue
                fuel_hist.setdefault(code, []).append((iso, price))
        history[fuel] = fuel_hist

        dates = sorted({d for d, _ in series})
        if not dates:
            continue
        last, prev = dates[-1].isoformat(), (dates[-2].isoformat() if len(dates) > 1 else None)
        by_date: dict[str, dict[str, float]] = {}
        for code, points in fuel_hist.items():
            for iso, price in points:
                by_date.setdefault(iso, {})[code] = price
        latest[fuel] = {
            "surveyDate": last,
            "prices": by_date.get(last, {}),
            "prevSurveyDate": prev,
            "prevPrices": by_date.get(prev, {}) if prev else {},
        }

    DATA_DIR.mkdir(exist_ok=True)
    (DATA_DIR / "history").mkdir(exist_ok=True)

    (DATA_DIR / "latest.json").write_text(
        json.dumps({"fuels": latest}, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )

    national = {fuel: hist.get("00", []) for fuel, hist in history.items()}
    (DATA_DIR / "history" / "national.json").write_text(
        json.dumps(national, ensure_ascii=False, separators=(",", ":")), encoding="utf-8"
    )
    for code in PREF_CODE.values():
        pref = {fuel: hist.get(code, []) for fuel, hist in history.items()}
        (DATA_DIR / "history" / f"{code}.json").write_text(
            json.dumps(pref, ensure_ascii=False, separators=(",", ":")), encoding="utf-8"
        )

    survey_dates = {f: v["surveyDate"] for f, v in latest.items()}
    (DATA_DIR / "meta.json").write_text(
        json.dumps(
            {
                "surveyDates": survey_dates,
                "source": "資源エネルギー庁「石油製品小売市況調査」",
                "sourceUrl": "https://www.enecho.meti.go.jp/statistics/petroleum_and_lpgas/pl007/",
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    print(f"OK: latest={survey_dates}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("usage: build_data.py <s5.xlsx>", file=sys.stderr)
        sys.exit(1)
    build(Path(sys.argv[1]))
